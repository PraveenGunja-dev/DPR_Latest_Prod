import openpyxl

wb = openpyxl.load_workbook(r'd:\DPR\Digitalized_DPR_Prod\data\DPR BESS _11 (1).xlsx', read_only=True, data_only=True)
with open('excel_headers.txt', 'w', encoding='utf-8') as f:
    f.write('Sheets: ' + ', '.join(wb.sheetnames) + '\n')
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        f.write(f'\n--- Sheet: {sheet} ---\n')
        # fetch first 10 rows
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
            # convert tuple to list of strings
            row_vals = [str(x) if x is not None else '' for x in row]
            # check if there's enough data to represent headers
            if len([x for x in row_vals if x]) > 3:
                f.write(f'Row {i+1}: ' + ' | '.join(row_vals[:20]) + '\n')
