import io
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_drone_excel(data: List[Dict[str, Any]], report_date: str) -> io.BytesIO:
    """
    Generate an Excel workbook with 4 sheets matching the exact categories from Spectra APIs:
    1. Construction Variation (block_progress)
    2. Inverter Variation (inverter_progress)
    3. Robot Variation (robot_progress)
    4. AC Work Variation (ac_work_progress)
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Common styles
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill("solid", fgColor="DDEBF7")  # Light blue
    drone_fill = PatternFill("solid", fgColor="FCE4D6")   # Light orange for drone
    dpr_fill = PatternFill("solid", fgColor="E2EFDA")     # Light green for dpr
    var_fill = PatternFill("solid", fgColor="FFF2CC")     # Light yellow for variance
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def apply_header_style(cell, fill_color=header_fill):
        cell.font = header_font
        cell.fill = fill_color
        cell.border = thin_border
        cell.alignment = center_align

    def apply_data_style(cell):
        cell.border = thin_border
        cell.alignment = center_align

    def build_variation_sheet(sheet_name: str, items: List[Dict[str, Any]]):
        ws = wb.create_sheet(sheet_name)
        
        if not items:
            ws.cell(row=1, column=1, value="No data available for this category.")
            ws.column_dimensions['A'].width = 30
            return

        # Collect all unique blocks across all items for this category
        all_blocks = set()
        for item in items:
            for b in item.get("block_breakdown", []):
                block_name = b.get("block")
                if block_name:
                    all_blocks.add(block_name)
        sorted_blocks = sorted(list(all_blocks))

        # Headers
        ws.merge_cells('A1:A2')
        ws['A1'] = "Block"
        apply_header_style(ws['A1'])
        apply_header_style(ws['A2'])

        col_idx = 2
        for item in items:
            act_name = item.get("activity") or "Activity"
            # Merge Top Header (3 columns per activity)
            start_col = get_column_letter(col_idx)
            end_col = get_column_letter(col_idx + 2)
            ws.merge_cells(f"{start_col}1:{end_col}1")
            ws[f"{start_col}1"] = act_name
            apply_header_style(ws[f"{start_col}1"])
            
            # Subheaders
            ws.cell(row=2, column=col_idx, value="DRONE")
            apply_header_style(ws.cell(row=2, column=col_idx), drone_fill)
            
            ws.cell(row=2, column=col_idx+1, value="DPR")
            apply_header_style(ws.cell(row=2, column=col_idx+1), dpr_fill)
            
            ws.cell(row=2, column=col_idx+2, value="VARIATION")
            apply_header_style(ws.cell(row=2, column=col_idx+2), var_fill)
            
            col_idx += 3

        # Rows
        row_idx = 3
        for block in sorted_blocks:
            ws.cell(row=row_idx, column=1, value=block)
            apply_data_style(ws.cell(row=row_idx, column=1))
            
            c_idx = 2
            for item in items:
                # Find block breakdown for this specific block
                block_data = {"drone": 0, "dpr": 0, "variance": 0}
                for b in item.get("block_breakdown", []):
                    if b.get("block") == block:
                        block_data = {
                            "drone": b.get("drone_actual", 0),
                            "dpr": b.get("dpr_actual", 0),
                            "variance": b.get("variance", 0)
                        }
                        break
                
                c_drone = ws.cell(row=row_idx, column=c_idx, value=block_data["drone"])
                apply_data_style(c_drone)
                
                c_dpr = ws.cell(row=row_idx, column=c_idx+1, value=block_data["dpr"])
                apply_data_style(c_dpr)
                
                c_var = ws.cell(row=row_idx, column=c_idx+2, value=block_data["variance"])
                apply_data_style(c_var)
                
                c_idx += 3
            row_idx += 1

        # Adjust column widths
        for col in range(1, col_idx):
            col_letter = get_column_letter(col)
            if col == 1:
                ws.column_dimensions[col_letter].width = 18
            else:
                ws.column_dimensions[col_letter].width = 15

    # Filter items based on spectra_api
    const_items = [item for item in data if item.get("spectra_api") == "block_progress"]
    inv_items = [item for item in data if item.get("spectra_api") == "inverter_progress"]
    robot_items = [item for item in data if item.get("spectra_api") == "robot_progress"]
    ac_items = [item for item in data if item.get("spectra_api") == "ac_work_progress"]

    # Build the sheets dynamically
    build_variation_sheet("Construction Variation", const_items)
    build_variation_sheet("Inverter Variation", inv_items)
    build_variation_sheet("Robot Variation", robot_items)
    build_variation_sheet("AC Work Variation", ac_items)

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
