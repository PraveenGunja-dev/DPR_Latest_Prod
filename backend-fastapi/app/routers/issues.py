# app/routers/issues.py
"""
Issues router – CRUD + stats.
Direct port of Express routes/issues.js
"""

import logging
from typing import Optional, Any
import json
from io import BytesIO
import pandas as pd

from fastapi import APIRouter, Depends, HTTPException, Query

from app.auth.dependencies import get_current_user
from app.database import get_db, PoolWrapper
from app.routers.project_utils import resolve_project_id

logger = logging.getLogger("adani-flow.issues")

router = APIRouter(prefix="/api/issues", tags=["Issues"])

ALLOWED_ROLES = {"site pm", "pmag", "super admin", "supervisor"}


def _check_pm_or_admin(user: dict[str, Any]):
    role = (user.get("role") or "").lower()
    if role not in ALLOWED_ROLES:
        raise HTTPException(403, detail={"error": "Access denied. Site PM, Supervisor, or Admin required."})


@router.get("")
async def get_issues(
    status: Optional[str] = None,
    priority: Optional[str] = None,
    project_id: Optional[str] = None,
    issue_type: Optional[str] = None,
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    pool: PoolWrapper = Depends(get_db),
    current_user: dict[str, Any] = Depends(get_current_user),
):
    _check_pm_or_admin(current_user)

    conditions = ["1=1"]
    params: list[Any] = []
    idx: int = 1

    if status:
        conditions.append(f"il.status = ${idx}"); params.append(status); idx += 1
    if priority:
        conditions.append(f"il.priority = ${idx}"); params.append(priority); idx += 1
    if project_id:
        project_object_id = await resolve_project_id(project_id, pool)
        conditions.append(f"il.project_id = ${idx}"); params.append(project_object_id); idx += 1
    if issue_type:
        conditions.append(f"il.issue_type = ${idx}"); params.append(issue_type); idx += 1

    where = " AND ".join(conditions)

    query = f"""
        SELECT il.*, u1.name as created_by_name, u1.email as created_by_email,
               u2.name as assigned_to_name, u3.name as resolved_by_name,
               COALESCE(p.name, p6."Name", 'No Project') as project_name
        FROM issue_logs il
        LEFT JOIN users u1 ON il.created_by = u1.user_id
        LEFT JOIN users u2 ON il.assigned_to = u2.user_id
        LEFT JOIN users u3 ON il.resolved_by = u3.user_id
        LEFT JOIN projects p ON il.project_id = p.object_id
        LEFT JOIN p6_projects p6 ON il.project_id = p6."ObjectId"
        WHERE {where}
        ORDER BY CASE il.priority WHEN 'critical' THEN 1 WHEN 'high' THEN 2 WHEN 'medium' THEN 3 WHEN 'low' THEN 4 END,
                 il.created_at DESC
        LIMIT ${idx} OFFSET ${idx + 1}
    """
    params.extend([limit, offset])

    rows = await pool.fetch(query, *params)

    # Count
    count_query = f"SELECT COUNT(*) FROM issue_logs il WHERE {where}"
    count_row = await pool.fetchval(count_query, *params[0:idx-1])

    return {
        "success": True,
        "issues": [dict(r) for r in rows],
        "total": count_row,
        "limit": limit,
        "offset": offset,
    }


@router.get("/stats/summary")
async def get_issue_stats(
    pool: PoolWrapper = Depends(get_db),
    current_user: dict[str, Any] = Depends(get_current_user),
):
    _check_pm_or_admin(current_user)

    row = await pool.fetchrow("""
        SELECT
            COUNT(*) FILTER (WHERE status = 'open') as open_count,
            COUNT(*) FILTER (WHERE status = 'in_progress') as in_progress_count,
            COUNT(*) FILTER (WHERE status = 'resolved') as resolved_count,
            COUNT(*) FILTER (WHERE status = 'closed') as closed_count,
            COUNT(*) FILTER (WHERE priority = 'critical') as critical_count,
            COUNT(*) FILTER (WHERE priority = 'high') as high_count,
            COUNT(*) as total_count
        FROM issue_logs
    """)
    if not row:
        return {"success": True, "stats": {
            "open_count": 0, "in_progress_count": 0, "resolved_count": 0,
            "closed_count": 0, "critical_count": 0, "high_count": 0, "total_count": 0
        }}

    return {"success": True, "stats": dict(row)}


@router.get("/{issue_id}")
async def get_issue(
    issue_id: int,
    pool: PoolWrapper = Depends(get_db),
    current_user: dict[str, Any] = Depends(get_current_user),
):
    _check_pm_or_admin(current_user)

    row = await pool.fetchrow("""
        SELECT il.*, u1.name as created_by_name, u1.email as created_by_email,
               u2.name as assigned_to_name, u3.name as resolved_by_name,
               COALESCE(p.name, p6."Name", 'No Project') as project_name
        FROM issue_logs il
        LEFT JOIN users u1 ON il.created_by = u1.user_id
        LEFT JOIN users u2 ON il.assigned_to = u2.user_id
        LEFT JOIN users u3 ON il.resolved_by = u3.user_id
        LEFT JOIN projects p ON il.project_id = p.object_id
        LEFT JOIN p6_projects p6 ON il.project_id = p6."ObjectId"
        WHERE il.id = $1
    """, issue_id)

    if not row:
        raise HTTPException(404, detail={"error": "Issue not found"})
    return {"success": True, "issue": dict(row)}


@router.post("", status_code=201)
async def create_issue(
    body: dict[str, Any],
    pool: PoolWrapper = Depends(get_db),
    current_user: dict[str, Any] = Depends(get_current_user),
):
    title = body.get("title")
    description = body.get("description")
    if not title or not description:
        raise HTTPException(400, detail={"error": "Title and description are required"})

    project_id = await resolve_project_id(body.get("project_id"), pool)

    row = await pool.fetchrow("""
        INSERT INTO issue_logs (project_id, entry_id, sheet_type, issue_type, title, description, priority, status, created_by, assigned_to)
        VALUES ($1, $2, $3, $4, $5, $6, $7, 'open', $8, $9)
        RETURNING *
    """,
        project_id, body.get("entry_id"), body.get("sheet_type"),
        body.get("issue_type", "general"), title, description,
        body.get("priority", "medium"), current_user["userId"],
        body.get("assigned_to"),
    )
    return {"success": True, "message": "Issue created successfully", "issue": dict(row)}


@router.put("/{issue_id}")
async def update_issue(
    issue_id: int,
    body: dict[str, Any],
    pool: PoolWrapper = Depends(get_db),
    current_user: dict[str, Any] = Depends(get_current_user),
):
    _check_pm_or_admin(current_user)

    updates: list[str] = []
    params: list[Any] = []
    idx: int = 1

    for field in ["title", "description", "issue_type", "priority", "status", "assigned_to", "resolution_notes"]:
        val = body.get(field)
        if val is not None:
            updates.append(f"{field} = ${idx}")
            params.append(val)
            idx += 1
            if field == "status" and val in ("resolved", "closed"):
                updates.append(f"resolved_by = ${idx}")
                params.append(current_user["userId"])
                idx += 1
                updates.append("resolved_at = CURRENT_TIMESTAMP")

    if not updates:
        raise HTTPException(400, detail={"error": "No fields to update"})

    params.append(issue_id)
    row = await pool.fetchrow(
        f"UPDATE issue_logs SET {', '.join(updates)} WHERE id = ${idx} RETURNING *",
        *params,
    )
    if not row:
        raise HTTPException(404, detail={"error": "Issue not found"})
    return {"success": True, "message": "Issue updated successfully", "issue": dict(row)}


@router.delete("/{issue_id}")
async def delete_issue(
    issue_id: int,
    pool: PoolWrapper = Depends(get_db),
    current_user: dict[str, Any] = Depends(get_current_user),
):
    _check_pm_or_admin(current_user)

    row = await pool.fetchrow("DELETE FROM issue_logs WHERE id = $1 RETURNING *", issue_id)
    if not row:
        raise HTTPException(404, detail={"error": "Issue not found"})
    return {"success": True, "message": "Issue deleted successfully"}


@router.post("/send-delay-alerts")
async def send_delay_alerts(
    pool: PoolWrapper = Depends(get_db),
    current_user: dict[str, Any] = Depends(get_current_user),
):
    _check_pm_or_admin(current_user)

    import re

    query = """
        SELECT sa.*, 
               COALESCE(p.name, p6."Name", 'No Project') as project_name,
               GREATEST(0, EXTRACT(DAY FROM (CURRENT_DATE - sa.planned_finish))) as delay_days
        FROM solar_activities sa
        LEFT JOIN projects p ON sa.project_object_id = p.object_id
        LEFT JOIN p6_projects p6 ON sa.project_object_id = p6."ObjectId"
        WHERE sa.status != 'Completed' 
          AND sa.planned_finish < CURRENT_DATE
          AND p.project_type ILIKE '%wind%'
        ORDER BY project_name, sa.planned_finish ASC
    """
    rows = await pool.fetch(query)

    from collections import defaultdict
    projects = defaultdict(list)
    
    def extract_activity_group(name: str, wbs_name: str) -> str:
        m = re.match(r'WTG\d+-(\w+)-', name or '', re.IGNORECASE)
        if m:
            code = m.group(1).upper()
            group_map = {'CW': 'CW', 'EL': 'EL', 'ELW': 'EL', 'TC': 'TC', 'ER': 'ER', 'ME': 'ME'}
            return group_map.get(code, code)
        wbs = (wbs_name or '').upper()
        if 'CIVIL' in wbs or 'CIVL' in wbs: return 'CW'
        if 'ELECTRIC' in wbs: return 'EL'
        if 'TESTING' in wbs or 'COMMISSION' in wbs: return 'TC'
        if 'ERECTION' in wbs: return 'ER'
        if 'EHV' in wbs or 'LINE' in wbs: return 'LINE'
        if 'PSS' in wbs: return 'PSS'
        if 'ENGINEER' in wbs: return 'ENG'
        return ''

    for r in rows:
        name = r["name"] or ""
        wbs_name = r["wbs_name"] or ""
        group = extract_activity_group(name, wbs_name)
        
        projects[r["project_name"]].append({
            "ACTIVITY ID": r["activity_id"],
            "DESCRIPTION": name,
            "GROUP": group,
            "LOCATION / WBS": wbs_name,
            "PLANNED FINISH": r["planned_finish"].strftime("%d %b %Y") if r["planned_finish"] else "-",
            "STATUS": r["status"]
        })

    excel_data = []
    
    for project_name, issues in projects.items():
        if not issues:
            continue
        excel_data.append({
            "ACTIVITY ID": f"Project: {project_name}",
            "DESCRIPTION": "",
            "GROUP": "",
            "LOCATION / WBS": "",
            "PLANNED FINISH": "",
            "STATUS": ""
        })
        for issue in issues:
            excel_data.append({
                "ACTIVITY ID": issue["ACTIVITY ID"],
                "DESCRIPTION": issue["DESCRIPTION"],
                "GROUP": issue["GROUP"],
                "LOCATION / WBS": issue["LOCATION / WBS"],
                "PLANNED FINISH": issue["PLANNED FINISH"],
                "STATUS": issue["STATUS"]
            })
        excel_data.append({
            "ACTIVITY ID": "", "DESCRIPTION": "", "GROUP": "", "LOCATION / WBS": "", "PLANNED FINISH": "", "STATUS": ""
        })

    df = pd.DataFrame(excel_data)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        if df.empty:
            df = pd.DataFrame([{"Message": "No delayed activities found."}])
            df.to_excel(writer, sheet_name="Delayed Activities", index=False)
        else:
            df.to_excel(writer, sheet_name="Delayed Activities", index=False)
            
            workbook = writer.book
            worksheet = writer.sheets["Delayed Activities"]
            from openpyxl.styles import Font, PatternFill
            bold_font = Font(bold=True)
            fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            for row_idx, row in enumerate(excel_data, start=2):
                if str(row["ACTIVITY ID"]).startswith("Project:"):
                    for col_idx in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.font = bold_font
                        cell.fill = fill

    excel_bytes = output.getvalue()

    from app.services.email_service import send_delay_alerts_email
    await send_delay_alerts_email(
        to_email="praveen.gunja@adani.com",
        sender_name=current_user.get("name", "PMAG Admin"),
        excel_bytes=excel_bytes
    )

    return {"success": True, "message": "Delay alerts sent successfully"}
